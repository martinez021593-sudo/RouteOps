
import csv
import io
import json
import math
import os
import secrets
import threading
import time as time_module
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, date, time, timezone
from functools import wraps
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from flask import (
    Flask, Response, abort, flash, jsonify, redirect, render_template,
    request, send_from_directory, session, url_for
)
from openpyxl import load_workbook
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
from PIL import Image, ImageOps
from db_layer import get_db, init_schema, INTEGRITY_ERRORS, database_backend
from intake_engine import parse_label_text, extract_label_data

BASE_DIR = Path(__file__).resolve().parent


def load_env_file():
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip())


load_env_file()

UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
TIMEZONE_NAME = os.environ.get("ROUTEOPS_TIMEZONE", "Europe/Madrid")
try:
    APP_TZ = ZoneInfo(TIMEZONE_NAME)
except Exception:
    APP_TZ = timezone.utc

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "routeops-v030-local-" + secrets.token_hex(16))
app.config["MAX_CONTENT_LENGTH"] = 12 * 1024 * 1024
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "0").strip().lower() in {"1","true","yes"}
app.permanent_session_lifetime = 60 * 60 * 24 * 7
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)
ALLOWED_IMAGE_EXT = {"png", "jpg", "jpeg", "webp"}


# V0.3.1.4 — background OCR queue.
# One Gunicorn worker is used by the Render pilot, while OCR is network-bound, so two background
# workers allow continuous scanning without blocking the camera UI.
_OCR_WORKERS = max(1, min(4, int(os.environ.get("INTAKE_OCR_WORKERS", "2"))))
OCR_EXECUTOR = ThreadPoolExecutor(max_workers=_OCR_WORKERS, thread_name_prefix="routeops-ocr")
OCR_SCHEDULED = set()
OCR_SCHEDULE_LOCK = threading.Lock()
OCR_DEBUG_CACHE = {}
OCR_DEBUG_LOCK = threading.Lock()
ROUTE_REOPT_TIMERS = {}
ROUTE_REOPT_LOCK = threading.Lock()


def now_local():
    return datetime.now(APP_TZ)


def now_iso():
    return now_local().isoformat(timespec="seconds")


def init_db():
    db = get_db()
    init_schema(db)
    if db.execute("SELECT COUNT(*) c FROM organizations").fetchone()["c"] == 0:
        depot_lat = _float_env("DEPOT_LAT")
        depot_lon = _float_env("DEPOT_LON")
        db.execute(
            "INSERT INTO organizations(name,timezone,currency,depot_lat,depot_lon,created_at) VALUES(?,?,?,?,?,?)",
            ("RouteOps Demo", TIMEZONE_NAME, "EUR", depot_lat, depot_lon, now_iso())
        )
        db.commit()

    org_id = db.execute("SELECT id FROM organizations ORDER BY id LIMIT 1").fetchone()["id"]

    if db.execute("SELECT COUNT(*) c FROM drivers").fetchone()["c"] == 0:
        for name in ("Carlos", "Juan", "Miguel"):
            db.execute(
                "INSERT INTO drivers(organization_id,name,email,pay_per_delivery,created_at) VALUES(?,?,?,?,?)",
                (org_id, name, f"{name.lower()}@routeops.local", 0.85, now_iso())
            )
        db.commit()

    if db.execute("SELECT COUNT(*) c FROM users").fetchone()["c"] == 0:
        db.execute(
            "INSERT INTO users(organization_id,email,username,password_hash,role) VALUES(?,?,?,?,?)",
            (org_id, os.environ.get("BOOTSTRAP_ADMIN_EMAIL", "admin@routeops.local"), "admin", generate_password_hash(os.environ.get("BOOTSTRAP_ADMIN_PASSWORD", "demo123")), "admin")
        )
        for d in db.execute("SELECT * FROM drivers WHERE organization_id=? ORDER BY id", (org_id,)):
            db.execute(
                "INSERT INTO users(organization_id,email,username,password_hash,role,driver_id) VALUES(?,?,?,?,?,?)",
                (org_id, d["email"], d["name"].lower(), generate_password_hash(os.environ.get("BOOTSTRAP_DRIVER_PASSWORD", "1234")), "driver", d["id"])
            )
        db.commit()

    if os.environ.get("SEED_DEMO_DATA", "1").strip().lower() not in {"0","false","no"}:
        today_s = date.today().isoformat()
        wd = db.execute(
            "SELECT * FROM work_days WHERE organization_id=? AND work_date=? ORDER BY id LIMIT 1",
            (org_id, today_s)
        ).fetchone()
        if not wd:
            cur = db.execute(
                "INSERT INTO work_days(organization_id,work_date,name,status,created_at,activated_at) VALUES(?,?,?,?,?,?)",
                (org_id, today_s, "Jornada Demo", "active", now_iso(), now_iso())
            )
            wd_id = cur.lastrowid
            seed_demo_packages(db, org_id, wd_id)
            db.commit()
    db.close()



def _float_env(key):
    val = os.environ.get(key, "").strip()
    try:
        return float(val) if val else None
    except Exception:
        return None


def seed_demo_packages(db, org_id, work_day_id):
    centers = {
        "Carlos": (40.4168, -3.7038),
        "Juan": (40.4380, -3.6900),
        "Miguel": (40.4010, -3.7150),
    }
    drivers = {r["name"]: r["id"] for r in db.execute(
        "SELECT * FROM drivers WHERE organization_id=?", (org_id,)
    )}
    idx = 1
    for name, (clat, clon) in centers.items():
        for n in range(10):
            angle = (n / 10) * math.tau + drivers[name]
            radius = 0.007 + (n % 4) * 0.0022
            lat = clat + math.sin(angle) * radius
            lon = clon + math.cos(angle) * radius
            db.execute(
                """
                INSERT INTO packages(
                    organization_id,work_day_id,code,barcode,recipient_name,phone,address,
                    lat,lon,driver_id,status,created_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    org_id, work_day_id, f"PK{idx:05d}", f"843700{idx:06d}",
                    f"Cliente {idx}", "", f"Parada demo {idx}, Madrid",
                    lat, lon, drivers[name], "pending", now_iso()
                )
            )
            idx += 1



def intake_carrier_counts(db, work_day_id, driver_id=None):
    params = [work_day_id]
    where_driver = ""
    if driver_id:
        where_driver = " AND driver_id=?"
        params.append(driver_id)
    rows = db.execute(
        f"""
        SELECT COALESCE(NULLIF(carrier,''),'unknown') carrier, COUNT(*) total,
               COALESCE(SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END),0) delivered
        FROM packages
        WHERE work_day_id=? {where_driver}
        GROUP BY COALESCE(NULLIF(carrier,''),'unknown')
        ORDER BY total DESC
        """, tuple(params)
    ).fetchall()
    return [dict(r) for r in rows]


def reoptimize_after_intake(db, wd, driver_id, org_id=None):
    if os.environ.get("AUTO_OPTIMIZE_INTAKE", "1").strip().lower() in {"0","false","no"}:
        return {"updated": False, "reason": "Auto-optimización desactivada"}
    org_id = int(org_id or get_org_id())
    org = db.execute("SELECT * FROM organizations WHERE id=?", (org_id,)).fetchone()
    try:
        count, km, mins, provider_name, note = optimize_driver_route(db, org, wd, driver_id, "local")
        return {"updated": True, "stops": count, "km": round(km,1), "minutes": mins, "provider": provider_name, "note": note}
    except Exception as exc:
        return {"updated": False, "reason": str(exc)}

def login_required(role=None):
    def deco(fn):
        @wraps(fn)
        def wrapped(*args, **kwargs):
            if not session.get("user_id"):
                return redirect(url_for("login"))
            if role and session.get("role") != role:
                flash("No tienes permisos para esa sección.", "error")
                return redirect(url_for("home"))
            return fn(*args, **kwargs)
        return wrapped
    return deco


def get_org_id():
    return int(session.get("organization_id") or 1)


def get_workday(db, work_day_id, org_id=None):
    org_id = org_id or get_org_id()
    row = db.execute(
        "SELECT * FROM work_days WHERE id=? AND organization_id=?",
        (work_day_id, org_id)
    ).fetchone()
    if not row:
        abort(404)
    return row


def active_workday_for_driver(db, org_id):
    return db.execute(
        """
        SELECT * FROM work_days
        WHERE organization_id=? AND status='active'
        ORDER BY CASE WHEN work_date=? THEN 0 ELSE 1 END, work_date DESC, id DESC
        LIMIT 1
        """,
        (org_id, date.today().isoformat())
    ).fetchone()


def haversine(a, b):
    lat1, lon1 = a
    lat2, lon2 = b
    r = 6371.0088
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    h = math.sin(dphi/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dlambda/2)**2
    return 2 * r * math.asin(math.sqrt(h))


def route_distance(points):
    return sum(haversine(points[i], points[i+1]) for i in range(len(points)-1)) if len(points) > 1 else 0


def nearest_neighbor_order(rows, start=None):
    valid = [dict(r) for r in rows if r["lat"] is not None and r["lon"] is not None]
    if not valid:
        return []
    current = start or (
        sum(r["lat"] for r in valid)/len(valid),
        sum(r["lon"] for r in valid)/len(valid)
    )
    remaining = valid[:]
    ordered = []
    while remaining:
        nxt = min(remaining, key=lambda r: haversine(current, (r["lat"], r["lon"])))
        ordered.append(nxt)
        current = (nxt["lat"], nxt["lon"])
        remaining.remove(nxt)

    improved = True
    loops = 0
    while improved and loops < 5 and len(ordered) < 220:
        improved = False
        loops += 1
        for i in range(1, len(ordered)-2):
            for j in range(i+1, min(len(ordered)-1, i+30)):
                a = (ordered[i-1]["lat"], ordered[i-1]["lon"])
                b = (ordered[i]["lat"], ordered[i]["lon"])
                c = (ordered[j]["lat"], ordered[j]["lon"])
                d = (ordered[j+1]["lat"], ordered[j+1]["lon"])
                if haversine(a,b)+haversine(c,d) > haversine(a,c)+haversine(b,d):
                    ordered[i:j+1] = list(reversed(ordered[i:j+1]))
                    improved = True
    return ordered


def parse_duration_seconds(value):
    if not value:
        return 0
    try:
        if isinstance(value, str) and value.endswith("s"):
            return float(value[:-1] or 0)
        return float(value)
    except Exception:
        return 0


def google_credentials():
    project = os.environ.get("GOOGLE_CLOUD_PROJECT", "").strip()
    cred_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    if not project or not cred_path:
        return None, "Falta GOOGLE_CLOUD_PROJECT o GOOGLE_APPLICATION_CREDENTIALS."
    path = Path(cred_path)
    if not path.is_absolute():
        path = BASE_DIR / path
    if not path.exists():
        return None, f"No existe el archivo de credenciales: {path}"
    try:
        from google.oauth2 import service_account
        from google.auth.transport.requests import Request as GoogleAuthRequest
        creds = service_account.Credentials.from_service_account_file(
            str(path), scopes=["https://www.googleapis.com/auth/cloud-platform"]
        )
        creds.refresh(GoogleAuthRequest())
        return (project, creds.token), None
    except Exception as exc:
        return None, f"Error OAuth Google: {exc}"


def google_route_optimize(rows, org, work_day):
    auth, err = google_credentials()
    if err:
        raise RuntimeError(err)
    project, token = auth

    valid = [dict(r) for r in rows if r["lat"] is not None and r["lon"] is not None]
    if not valid:
        raise RuntimeError("No hay paradas geocodificadas.")

    depot_lat = org["depot_lat"]
    depot_lon = org["depot_lon"]
    if depot_lat is None or depot_lon is None:
        depot_lat = sum(r["lat"] for r in valid) / len(valid)
        depot_lon = sum(r["lon"] for r in valid) / len(valid)

    tzname = org["timezone"] or TIMEZONE_NAME
    try:
        tz = ZoneInfo(tzname)
    except Exception:
        tz = APP_TZ

    d = date.fromisoformat(work_day["work_date"])
    start_local = datetime.combine(d, time(6, 0), tzinfo=tz)
    end_local = datetime.combine(d, time(23, 30), tzinfo=tz)
    start_utc = start_local.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
    end_utc = end_local.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")

    shipments = []
    for r in valid:
        shipments.append({
            "deliveries": [{
                "arrivalWaypoint": {
                    "location": {
                        "latLng": {
                            "latitude": float(r["lat"]),
                            "longitude": float(r["lon"])
                        }
                    }
                },
                "duration": "180s"
            }],
            "label": r["code"]
        })

    waypoint = {
        "location": {
            "latLng": {"latitude": float(depot_lat), "longitude": float(depot_lon)}
        }
    }
    payload = {
        "timeout": "20s",
        "model": {
            "shipments": shipments,
            "vehicles": [{
                "startWaypoint": waypoint,
                "endWaypoint": waypoint,
                "costPerKilometer": 1.0,
                "startTimeWindows": [{"startTime": start_utc, "endTime": start_utc}],
                "endTimeWindows": [{"startTime": start_utc, "endTime": end_utc}]
            }],
            "globalStartTime": start_utc,
            "globalEndTime": end_utc
        }
    }

    url = f"https://routeoptimization.googleapis.com/v1/projects/{project}:optimizeTours"
    resp = requests.post(
        url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json=payload,
        timeout=70
    )
    if not resp.ok:
        try:
            detail = resp.json()
        except Exception:
            detail = resp.text[:800]
        raise RuntimeError(f"Google Route Optimization HTTP {resp.status_code}: {detail}")

    data = resp.json()
    if data.get("skippedShipments"):
        skipped = len(data["skippedShipments"])
    else:
        skipped = 0
    routes = data.get("routes") or []
    if not routes:
        raise RuntimeError("Google no devolvió una ruta.")

    route = routes[0]
    visits = route.get("visits") or []
    ordered = []
    used = set()
    for visit in visits:
        # En entregas, isPickup=false puede omitirse en JSON.
        if visit.get("isPickup", False):
            continue
        idx = int(visit.get("shipmentIndex", 0))
        if 0 <= idx < len(valid) and idx not in used:
            ordered.append(valid[idx])
            used.add(idx)

    # Defensa: si Google omite visitas o devuelve alguna parada sin secuenciar, la añadimos al final.
    for idx, r in enumerate(valid):
        if idx not in used:
            ordered.append(r)

    metrics = route.get("metrics") or {}
    km = float(metrics.get("travelDistanceMeters") or 0) / 1000.0
    total_seconds = parse_duration_seconds(metrics.get("totalDuration"))
    mins = max(1, int(round(total_seconds / 60))) if total_seconds else 0
    note = f"Google Route Optimization · {len(ordered)} visitas"
    if skipped:
        note += f" · {skipped} omitidas por Google"
    return ordered, km, mins, note


def optimize_driver_route(db, org, work_day, driver_id, provider):
    rows = db.execute(
        """
        SELECT * FROM packages
        WHERE work_day_id=? AND driver_id=? AND status IN ('pending','failed')
        ORDER BY id
        """,
        (work_day["id"], driver_id)
    ).fetchall()
    if not rows:
        raise RuntimeError("El repartidor no tiene paquetes pendientes.")

    geocoded = [r for r in rows if r["lat"] is not None and r["lon"] is not None]
    if not geocoded:
        raise RuntimeError("El repartidor no tiene paradas con coordenadas.")

    if provider == "google":
        ordered, km, mins, note = google_route_optimize(geocoded, org, work_day)
        provider_name = "google"
    else:
        start = None
        if org["depot_lat"] is not None and org["depot_lon"] is not None:
            start = (org["depot_lat"], org["depot_lon"])
        ordered = nearest_neighbor_order(geocoded, start=start)
        pts = []
        if start:
            pts.append(start)
        pts.extend((r["lat"], r["lon"]) for r in ordered)
        if start:
            pts.append(start)
        km = route_distance(pts)
        mins = int((km / 24) * 60 + len(ordered) * 3)
        note = "Local NN + 2-opt · distancia geográfica aproximada"
        provider_name = "local"

    for seq, r in enumerate(ordered, start=1):
        db.execute("UPDATE packages SET sequence=? WHERE id=?", (seq, r["id"]))

    db.execute(
        """
        INSERT INTO route_runs(
            organization_id,work_day_id,driver_id,created_at,provider,
            total_distance_km,estimated_minutes,stop_count,status,provider_note
        ) VALUES(?,?,?,?,?,?,?,?,?,?)
        """,
        (
            org["id"], work_day["id"], driver_id, now_iso(), provider_name,
            km, mins, len(ordered), "published", note
        )
    )
    return len(ordered), km, mins, provider_name, note


def geocode_google(address):
    key = os.environ.get("GOOGLE_MAPS_API_KEY", "").strip()
    if not key:
        return None
    try:
        resp = requests.get(
            "https://maps.googleapis.com/maps/api/geocode/json",
            params={"address": address, "key": key}, timeout=12
        )
        data = resp.json()
        if data.get("status") == "OK" and data.get("results"):
            loc = data["results"][0]["geometry"]["location"]
            return float(loc["lat"]), float(loc["lng"])
    except Exception:
        return None
    return None



def _json_load_list(value):
    if not value:
        return []
    try:
        data = json.loads(value)
        if isinstance(data, list):
            return [str(x) for x in data if str(x).strip()]
    except Exception:
        pass
    return [str(value)]


def _clean_job_result(result):
    # Never persist raw OCR text. Keep only operational fields needed by the authorized driver UI.
    allowed = {
        "package_id", "carrier", "carrier_reason", "tracking_code", "tracking_source", "barcode",
        "recipient_name", "address", "postal_code", "city", "route_zone", "route_code",
        "weight_kg", "quantity", "confidence", "intake_status", "profile", "missing_required",
        "detected_fields", "geocoded", "geocode_status", "ocr_confidence", "ocr_passes",
        "duplicate", "message",
    }
    return {k: v for k, v in result.items() if k in allowed}


def _remember_ocr_debug(job_id, text):
    if os.environ.get("INTAKE_OCR_DEBUG", "1").strip().lower() in {"0", "false", "no"}:
        return
    with OCR_DEBUG_LOCK:
        OCR_DEBUG_CACHE[int(job_id)] = (time_module.time(), (text or "")[:3200])
        # Keep only a short-lived in-memory diagnostic cache; never persist full OCR text.
        cutoff = time_module.time() - 600
        for key, (ts, _) in list(OCR_DEBUG_CACHE.items()):
            if ts < cutoff:
                OCR_DEBUG_CACHE.pop(key, None)


def _get_ocr_debug(job_id):
    with OCR_DEBUG_LOCK:
        item = OCR_DEBUG_CACHE.get(int(job_id))
        if not item:
            return ""
        if time_module.time() - item[0] > 600:
            OCR_DEBUG_CACHE.pop(int(job_id), None)
            return ""
        return item[1]


def _schedule_route_refresh(org_id, work_day_id, driver_id):
    if os.environ.get("AUTO_OPTIMIZE_INTAKE", "1").strip().lower() in {"0", "false", "no"}:
        return
    key = (int(org_id), int(work_day_id), int(driver_id))

    def run():
        try:
            db = get_db()
            wd = db.execute(
                "SELECT * FROM work_days WHERE id=? AND organization_id=?", (work_day_id, org_id)
            ).fetchone()
            org = db.execute("SELECT * FROM organizations WHERE id=?", (org_id,)).fetchone()
            if wd and org:
                try:
                    optimize_driver_route(db, org, wd, driver_id, "local")
                    db.commit()
                except Exception:
                    db.rollback()
            db.close()
        finally:
            with ROUTE_REOPT_LOCK:
                ROUTE_REOPT_TIMERS.pop(key, None)

    with ROUTE_REOPT_LOCK:
        old_timer = ROUTE_REOPT_TIMERS.get(key)
        if old_timer:
            old_timer.cancel()
        # Debounce: scanning a burst of labels results in one route refresh after the burst.
        timer = threading.Timer(1.8, run)
        timer.daemon = True
        ROUTE_REOPT_TIMERS[key] = timer
        timer.start()


def _process_intake_job(job_id):
    db = None
    try:
        db = get_db()
        job = db.execute("SELECT * FROM intake_jobs WHERE id=?", (job_id,)).fetchone()
        if not job or job["status"] not in ("queued", "processing"):
            return
        db.execute(
            "UPDATE intake_jobs SET status='processing',started_at=?,attempts=COALESCE(attempts,0)+1,error_text=NULL WHERE id=?",
            (now_iso(), job_id),
        )
        db.commit()

        image_data = job["image_data"]
        if isinstance(image_data, memoryview):
            image_data = image_data.tobytes()
        if not image_data:
            raise RuntimeError("La imagen de la etiqueta no está disponible")
        raw_codes = _json_load_list(job["raw_codes"])
        raw_code = raw_codes[0] if raw_codes else ""
        carrier_hint = (job["carrier_hint"] or "").strip().lower()
        if carrier_hint == "agencia":
            carrier_hint = "tipsa"
        if carrier_hint not in {"imile", "ecoscooting", "tipsa"}:
            carrier_hint = ""

        # High-accuracy OCR. If important fields are missing the engine automatically executes pass 2.
        parsed = extract_label_data(
            bytes(image_data), raw_code=raw_code, raw_codes=raw_codes, forced_carrier=carrier_hint
        )
        _remember_ocr_debug(job_id, parsed.pop("ocr_text", ""))

        tracking = (parsed.get("tracking_code") or "").strip()
        carrier = parsed.get("carrier") or "unknown"
        barcode = (parsed.get("barcode") or raw_code or "").strip()
        code = tracking or f"REVIEW-{job_id}"

        # Duplicate detection is based primarily on the carrier-selected tracking, not whichever
        # barcode happened to be seen first by the browser.
        existing = None
        if tracking:
            existing = db.execute(
                """
                SELECT * FROM packages
                WHERE work_day_id=? AND (tracking_code=? OR code=?)
                ORDER BY id LIMIT 1
                """, (job["work_day_id"], tracking, tracking)
            ).fetchone()
        if existing:
            result = {
                "package_id": existing["id"], "carrier": existing["carrier"] or carrier,
                "tracking_code": existing["tracking_code"] or existing["code"],
                "tracking_source": parsed.get("tracking_source") or "",
                "barcode": barcode, "recipient_name": existing["recipient_name"] or "",
                "address": existing["address"] or "", "postal_code": existing["postal_code"] or "",
                "city": existing["city"] or "", "route_zone": existing["route_zone"] or "",
                "route_code": existing["route_code"] or "", "confidence": parsed.get("intake_confidence") or 0,
                "intake_status": existing["intake_status"] or "ready", "profile": parsed.get("profile") or "",
                "missing_required": parsed.get("missing_required") or [], "detected_fields": parsed.get("detected_fields") or {},
                "ocr_confidence": parsed.get("ocr_confidence"), "ocr_passes": parsed.get("ocr_passes") or 1,
                "duplicate": True, "message": "Este tracking ya estaba registrado",
            }
            db.execute(
                "UPDATE intake_jobs SET status='duplicate',package_id=?,result_json=?,image_data=NULL,completed_at=? WHERE id=?",
                (existing["id"], json.dumps(_clean_job_result(result), ensure_ascii=False), now_iso(), job_id),
            )
            db.commit()
            return

        address = (parsed.get("address") or "").strip()
        lat = lon = None
        geocode_status = "not_requested"
        if address:
            if os.environ.get("GOOGLE_MAPS_API_KEY", "").strip():
                loc = geocode_google(address)
                if loc:
                    lat, lon = loc
                    geocode_status = "ok"
                else:
                    geocode_status = "failed"
            else:
                geocode_status = "not_configured"

        cur = db.execute(
            """
            INSERT INTO packages(
                organization_id,work_day_id,code,barcode,carrier,tracking_code,postal_code,city,
                route_zone,route_code,weight_kg,quantity,intake_source,intake_driver_id,
                intake_scanned_at,intake_confidence,intake_status,raw_scan_code,
                tracking_source,ocr_confidence,ocr_passes,intake_job_id,
                recipient_name,phone,address,lat,lon,driver_id,status,created_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                job["organization_id"], job["work_day_id"], code, barcode, carrier, tracking or None,
                parsed.get("postal_code") or "", parsed.get("city") or "", parsed.get("route_zone") or "",
                parsed.get("route_code") or "", parsed.get("weight_kg"), parsed.get("quantity") or 1,
                "camera", job["driver_id"], now_iso(), parsed.get("intake_confidence") or 0,
                parsed.get("intake_status") or "review", json.dumps(raw_codes, ensure_ascii=False),
                parsed.get("tracking_source") or "", parsed.get("ocr_confidence"), parsed.get("ocr_passes") or 1,
                job_id, parsed.get("recipient_name") or "", parsed.get("phone") or "",
                address or "PENDIENTE DE REVISIÓN", lat, lon, job["driver_id"], "pending", now_iso(),
            ),
        )
        package_id = cur.lastrowid
        db.execute(
            """
            INSERT INTO intake_events(
                organization_id,work_day_id,driver_id,package_id,carrier,source,confidence,status,captured_at
            ) VALUES(?,?,?,?,?,?,?,?,?)
            """,
            (
                job["organization_id"], job["work_day_id"], job["driver_id"], package_id, carrier,
                "camera-background", parsed.get("intake_confidence") or 0,
                parsed.get("intake_status") or "review", now_iso(),
            ),
        )
        result = {
            "package_id": package_id, "carrier": carrier, "carrier_reason": parsed.get("carrier_reason"),
            "tracking_code": tracking, "tracking_source": parsed.get("tracking_source") or "", "barcode": barcode,
            "recipient_name": parsed.get("recipient_name") or "", "address": address,
            "postal_code": parsed.get("postal_code") or "", "city": parsed.get("city") or "",
            "route_zone": parsed.get("route_zone") or "", "route_code": parsed.get("route_code") or "",
            "weight_kg": parsed.get("weight_kg"), "quantity": parsed.get("quantity") or 1,
            "confidence": parsed.get("intake_confidence") or 0, "intake_status": parsed.get("intake_status") or "review",
            "profile": parsed.get("profile") or "generic_v2", "missing_required": parsed.get("missing_required") or [],
            "detected_fields": parsed.get("detected_fields") or {}, "geocoded": lat is not None and lon is not None,
            "geocode_status": geocode_status, "ocr_confidence": parsed.get("ocr_confidence"),
            "ocr_passes": parsed.get("ocr_passes") or 1, "duplicate": False,
        }
        db.execute(
            "UPDATE intake_jobs SET status='done',package_id=?,result_json=?,image_data=NULL,completed_at=? WHERE id=?",
            (package_id, json.dumps(_clean_job_result(result), ensure_ascii=False), now_iso(), job_id),
        )
        db.commit()
        if lat is not None and lon is not None:
            _schedule_route_refresh(job["organization_id"], job["work_day_id"], job["driver_id"])
    except Exception as exc:
        if db:
            try:
                db.rollback()
                db.execute(
                    "UPDATE intake_jobs SET status='error',error_text=?,image_data=NULL,completed_at=? WHERE id=?",
                    (str(exc)[:700], now_iso(), job_id),
                )
                db.commit()
            except Exception:
                pass
    finally:
        if db:
            try:
                db.close()
            except Exception:
                pass


def _intake_job_finished(job_id, future):
    with OCR_SCHEDULE_LOCK:
        OCR_SCHEDULED.discard(int(job_id))


def schedule_intake_job(job_id):
    job_id = int(job_id)
    with OCR_SCHEDULE_LOCK:
        if job_id in OCR_SCHEDULED:
            return False
        OCR_SCHEDULED.add(job_id)
    future = OCR_EXECUTOR.submit(_process_intake_job, job_id)
    future.add_done_callback(lambda f, jid=job_id: _intake_job_finished(jid, f))
    return True


def resume_intake_jobs(work_day_id, driver_id):
    db = get_db()
    rows = db.execute(
        """
        SELECT id FROM intake_jobs
        WHERE work_day_id=? AND driver_id=? AND status IN ('queued','processing')
        ORDER BY id LIMIT 40
        """, (work_day_id, driver_id)
    ).fetchall()
    db.close()
    for row in rows:
        schedule_intake_job(row["id"])


def read_import(file_storage):
    name = (file_storage.filename or "").lower()
    raw = file_storage.read()
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        rows = list(csv.DictReader(io.StringIO(text)))
    elif name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(x or "").strip() for x in values[0]]
        rows = [dict(zip(headers, row)) for row in values[1:]]
    else:
        raise ValueError("Formato no soportado. Usa CSV o XLSX.")

    aliases = {
        "code": ["codigo", "código", "code", "paquete", "package_code"],
        "barcode": ["barcode", "codigo_barras", "código_barras", "qr", "tracking"],
        "recipient": ["cliente", "destinatario", "recipient", "recipient_name", "nombre"],
        "phone": ["telefono", "teléfono", "phone", "movil", "móvil"],
        "address": ["direccion", "dirección", "address", "domicilio"],
        "driver": ["conductor", "repartidor", "driver"],
        "lat": ["lat", "latitude", "latitud"],
        "lon": ["lon", "lng", "longitude", "longitud"],
    }

    def pick(row, keys):
        normalized = {str(k).strip().lower(): v for k, v in row.items()}
        for k in keys:
            if k in normalized and normalized[k] not in (None, ""):
                return normalized[k]
        return ""

    items = []
    for row in rows:
        code = str(pick(row, aliases["code"])).strip()
        address = str(pick(row, aliases["address"])).strip()
        if not code or not address:
            continue
        lat_v, lon_v = pick(row, aliases["lat"]), pick(row, aliases["lon"])
        try:
            lat_v = float(lat_v) if lat_v != "" else None
        except Exception:
            lat_v = None
        try:
            lon_v = float(lon_v) if lon_v != "" else None
        except Exception:
            lon_v = None
        barcode = str(pick(row, aliases["barcode"])).strip() or code
        items.append({
            "code": code,
            "barcode": barcode,
            "recipient": str(pick(row, aliases["recipient"])).strip(),
            "phone": str(pick(row, aliases["phone"])).strip(),
            "address": address,
            "driver": str(pick(row, aliases["driver"])).strip(),
            "lat": lat_v,
            "lon": lon_v,
        })
    return items


@app.context_processor
def inject_globals():
    return {
        "today": date.today().isoformat(),
        "routeops_timezone": TIMEZONE_NAME
    }


@app.route("/")
def home():
    if not session.get("user_id"):
        return redirect(url_for("login"))
    return redirect(url_for("admin_dashboard") if session.get("role") == "admin" else url_for("driver_home"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        identity = request.form.get("identity", "").strip().lower()
        password = request.form.get("password", "")
        db = get_db()
        user = db.execute(
            """
            SELECT * FROM users
            WHERE lower(coalesce(email,''))=? OR lower(coalesce(username,''))=?
            ORDER BY id LIMIT 1
            """,
            (identity, identity)
        ).fetchone()
        db.close()
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session.update({
                "user_id": user["id"],
                "organization_id": user["organization_id"],
                "role": user["role"],
                "driver_id": user["driver_id"]
            })
            return redirect(url_for("home"))
        flash("Credenciales incorrectas.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/admin")
@login_required("admin")
def admin_dashboard():
    db = get_db()
    org_id = get_org_id()
    workdays = db.execute(
        "SELECT * FROM work_days WHERE organization_id=? ORDER BY work_date DESC,id DESC LIMIT 30",
        (org_id,)
    ).fetchall()
    work_day_id = request.args.get("work_day_id", type=int)
    if work_day_id:
        wd = get_workday(db, work_day_id, org_id)
    else:
        wd = db.execute(
            """
            SELECT * FROM work_days WHERE organization_id=?
            ORDER BY CASE WHEN status='active' THEN 0 ELSE 1 END, work_date DESC,id DESC LIMIT 1
            """, (org_id,)
        ).fetchone()

    stats = {"assigned": 0, "delivered": 0, "pending": 0, "failed": 0}
    drivers = []
    if wd:
        stats_row = db.execute(
            """
            SELECT COUNT(*) assigned,
                   COALESCE(SUM(status='delivered'),0) delivered,
                   COALESCE(SUM(status='pending'),0) pending,
                   COALESCE(SUM(status='failed'),0) failed
            FROM packages WHERE work_day_id=?
            """, (wd["id"],)
        ).fetchone()
        stats = dict(stats_row)
        drivers = db.execute(
            """
            SELECT d.*, COUNT(p.id) assigned,
                   COALESCE(SUM(p.status='delivered'),0) delivered,
                   COALESCE(SUM(p.status='failed'),0) failed,
                   COALESCE(SUM(p.status='pending'),0) pending
            FROM drivers d
            LEFT JOIN packages p ON p.driver_id=d.id AND p.work_day_id=?
            WHERE d.organization_id=? AND d.active=1
            GROUP BY d.id ORDER BY d.name
            """, (wd["id"], org_id)
        ).fetchall()
    db.close()
    return render_template("admin_dashboard.html", stats=stats, drivers=drivers, workdays=workdays, wd=wd)


@app.route("/admin/workdays")
@login_required("admin")
def workdays_page():
    db = get_db()
    org_id = get_org_id()
    workdays = db.execute(
        """
        SELECT w.*,
               COUNT(p.id) packages,
               COALESCE(SUM(p.status='delivered'),0) delivered,
               COALESCE(SUM(p.status='failed'),0) failed
        FROM work_days w
        LEFT JOIN packages p ON p.work_day_id=w.id
        WHERE w.organization_id=?
        GROUP BY w.id ORDER BY w.work_date DESC,w.id DESC
        """, (org_id,)
    ).fetchall()
    db.close()
    return render_template("workdays.html", workdays=workdays)


@app.route("/admin/workdays", methods=["POST"])
@login_required("admin")
def create_workday():
    work_date = request.form.get("work_date") or date.today().isoformat()
    name = request.form.get("name", "").strip() or "Jornada"
    org_id = get_org_id()
    db = get_db()
    try:
        cur = db.execute(
            "INSERT INTO work_days(organization_id,work_date,name,status,created_at) VALUES(?,?,?,?,?)",
            (org_id, work_date, name, "planning", now_iso())
        )
        db.commit()
        flash("Jornada creada.", "success")
        return redirect(url_for("workday_detail", work_day_id=cur.lastrowid))
    except INTEGRITY_ERRORS:
        flash("Ya existe una jornada con ese nombre en esa fecha.", "error")
        return redirect(url_for("workdays_page"))
    finally:
        db.close()


@app.route("/admin/workdays/<int:work_day_id>")
@login_required("admin")
def workday_detail(work_day_id):
    db = get_db()
    wd = get_workday(db, work_day_id)
    stats = db.execute(
        """
        SELECT COUNT(*) packages,
               COALESCE(SUM(status='delivered'),0) delivered,
               COALESCE(SUM(status='pending'),0) pending,
               COALESCE(SUM(status='failed'),0) failed,
               COALESCE(SUM(lat IS NULL OR lon IS NULL),0) missing_coords
        FROM packages WHERE work_day_id=?
        """, (work_day_id,)
    ).fetchone()
    drivers = db.execute(
        """
        SELECT d.*, COUNT(p.id) packages
        FROM drivers d LEFT JOIN packages p ON p.driver_id=d.id AND p.work_day_id=?
        WHERE d.organization_id=? AND d.active=1
        GROUP BY d.id ORDER BY d.name
        """, (work_day_id, get_org_id())
    ).fetchall()
    db.close()
    return render_template("workday_detail.html", wd=wd, stats=stats, drivers=drivers)


@app.route("/admin/workdays/<int:work_day_id>/status", methods=["POST"])
@login_required("admin")
def workday_status(work_day_id):
    status = request.form.get("status")
    if status not in {"planning", "active", "closed"}:
        abort(400)
    db = get_db()
    wd = get_workday(db, work_day_id)
    if status == "active":
        db.execute(
            "UPDATE work_days SET status='planning' WHERE organization_id=? AND status='active' AND id<>?",
            (get_org_id(), work_day_id)
        )
        db.execute(
            "UPDATE work_days SET status='active',activated_at=COALESCE(activated_at,?) WHERE id=?",
            (now_iso(), work_day_id)
        )
    elif status == "closed":
        db.execute("UPDATE work_days SET status='closed',closed_at=? WHERE id=?", (now_iso(), work_day_id))
    else:
        db.execute("UPDATE work_days SET status='planning' WHERE id=?", (work_day_id,))
    db.commit()
    db.close()
    flash(f"Jornada actualizada a {status}.", "success")
    return redirect(url_for("workday_detail", work_day_id=work_day_id))



@app.route("/admin/workdays/<int:work_day_id>/intake")
@login_required("admin")
def intake_admin_page(work_day_id):
    db = get_db()
    wd = get_workday(db, work_day_id)
    counts = intake_carrier_counts(db, work_day_id)
    by_driver = db.execute(
        """
        SELECT d.id,d.name,COUNT(p.id) total,
               COALESCE(SUM(CASE WHEN p.carrier='imile' THEN 1 ELSE 0 END),0) imile,
               COALESCE(SUM(CASE WHEN p.carrier='ecoscooting' THEN 1 ELSE 0 END),0) ecoscooting,
               COALESCE(SUM(CASE WHEN p.carrier IN ('tipsa','agencia') THEN 1 ELSE 0 END),0) tipsa,
               COALESCE(SUM(CASE WHEN p.intake_status='review' THEN 1 ELSE 0 END),0) review,
               COALESCE(SUM(CASE WHEN p.status='delivered' THEN 1 ELSE 0 END),0) delivered
        FROM drivers d LEFT JOIN packages p ON p.driver_id=d.id AND p.work_day_id=?
        WHERE d.organization_id=? AND d.active=1
        GROUP BY d.id,d.name ORDER BY d.name
        """, (work_day_id, get_org_id())
    ).fetchall()
    recent = db.execute(
        """
        SELECT p.id,p.code,p.tracking_code,p.carrier,p.recipient_name,p.address,p.postal_code,
               p.route_zone,p.route_code,p.intake_confidence,p.intake_status,p.intake_scanned_at,
               p.status,d.name driver_name,dd.name delivered_by
        FROM packages p
        LEFT JOIN drivers d ON d.id=p.driver_id
        LEFT JOIN drivers dd ON dd.id=p.delivered_by_driver_id
        WHERE p.work_day_id=? AND p.intake_source='camera'
        ORDER BY p.id DESC LIMIT 250
        """, (work_day_id,)
    ).fetchall()
    db.close()
    return render_template("intake_admin.html", wd=wd, counts=counts, by_driver=by_driver, recent=recent)

@app.route("/admin/workdays/<int:work_day_id>/packages")
@login_required("admin")
def packages_page(work_day_id):
    db = get_db()
    wd = get_workday(db, work_day_id)
    packages = db.execute(
        """
        SELECT p.*, d.name driver_name FROM packages p
        LEFT JOIN drivers d ON d.id=p.driver_id
        WHERE p.work_day_id=?
        ORDER BY CASE WHEN p.sequence IS NULL THEN 1 ELSE 0 END,p.sequence,p.id
        LIMIT 1500
        """, (work_day_id,)
    ).fetchall()
    drivers = db.execute(
        "SELECT * FROM drivers WHERE organization_id=? AND active=1 ORDER BY name",
        (get_org_id(),)
    ).fetchall()
    missing_coords = db.execute(
        "SELECT COUNT(*) c FROM packages WHERE work_day_id=? AND (lat IS NULL OR lon IS NULL)",
        (work_day_id,)
    ).fetchone()["c"]
    db.close()
    return render_template(
        "packages.html", wd=wd, packages=packages, drivers=drivers, missing_coords=missing_coords
    )


@app.route("/admin/workdays/<int:work_day_id>/packages/import", methods=["POST"])
@login_required("admin")
def import_packages(work_day_id):
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Selecciona un CSV o XLSX.", "error")
        return redirect(url_for("packages_page", work_day_id=work_day_id))
    try:
        items = read_import(f)
    except Exception as exc:
        flash(str(exc), "error")
        return redirect(url_for("packages_page", work_day_id=work_day_id))

    db = get_db()
    wd = get_workday(db, work_day_id)
    drivers = {
        r["name"].lower(): r["id"]
        for r in db.execute("SELECT * FROM drivers WHERE organization_id=?", (get_org_id(),))
    }
    created, duplicated = 0, 0
    for it in items:
        driver_id = drivers.get(it["driver"].lower()) if it["driver"] else None
        try:
            db.execute(
                """
                INSERT INTO packages(
                    organization_id,work_day_id,code,barcode,recipient_name,phone,address,
                    lat,lon,driver_id,status,created_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    get_org_id(), work_day_id, it["code"], it["barcode"], it["recipient"],
                    it["phone"], it["address"], it["lat"], it["lon"], driver_id,
                    "pending", now_iso()
                )
            )
            created += 1
        except INTEGRITY_ERRORS:
            duplicated += 1
    db.commit()
    db.close()
    flash(f"Importación: {created} nuevos · {duplicated} duplicados omitidos.", "success")
    return redirect(url_for("packages_page", work_day_id=work_day_id))


@app.route("/admin/workdays/<int:work_day_id>/packages/manual", methods=["POST"])
@login_required("admin")
def add_package_manual(work_day_id):
    code = request.form.get("code", "").strip()
    barcode = request.form.get("barcode", "").strip() or code
    address = request.form.get("address", "").strip()
    if not code or not address:
        flash("Código y dirección son obligatorios.", "error")
        return redirect(url_for("packages_page", work_day_id=work_day_id))
    driver_id = request.form.get("driver_id", type=int)
    lat = request.form.get("lat", type=float)
    lon = request.form.get("lon", type=float)
    db = get_db()
    get_workday(db, work_day_id)
    try:
        db.execute(
            """
            INSERT INTO packages(
                organization_id,work_day_id,code,barcode,recipient_name,phone,address,
                lat,lon,driver_id,status,created_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                get_org_id(), work_day_id, code, barcode,
                request.form.get("recipient_name", "").strip(),
                request.form.get("phone", "").strip(),
                address, lat, lon, driver_id, "pending", now_iso()
            )
        )
        db.commit()
        flash("Paquete añadido.", "success")
    except INTEGRITY_ERRORS:
        flash("Ese código ya existe en la jornada.", "error")
    finally:
        db.close()
    return redirect(url_for("packages_page", work_day_id=work_day_id))


@app.route("/admin/workdays/<int:work_day_id>/packages/<int:package_id>/assign", methods=["POST"])
@login_required("admin")
def assign_package(work_day_id, package_id):
    driver_id = request.form.get("driver_id", type=int)
    db = get_db()
    get_workday(db, work_day_id)
    db.execute(
        "UPDATE packages SET driver_id=?,sequence=NULL WHERE id=? AND work_day_id=?",
        (driver_id, package_id, work_day_id)
    )
    db.commit()
    db.close()
    return redirect(url_for("packages_page", work_day_id=work_day_id))


@app.route("/admin/workdays/<int:work_day_id>/geocode", methods=["POST"])
@login_required("admin")
def geocode_missing(work_day_id):
    if not os.environ.get("GOOGLE_MAPS_API_KEY", "").strip():
        flash("Configura GOOGLE_MAPS_API_KEY en .env.", "error")
        return redirect(url_for("packages_page", work_day_id=work_day_id))
    db = get_db()
    get_workday(db, work_day_id)
    rows = db.execute(
        "SELECT id,address FROM packages WHERE work_day_id=? AND (lat IS NULL OR lon IS NULL) LIMIT 300",
        (work_day_id,)
    ).fetchall()
    ok = 0
    for r in rows:
        loc = geocode_google(r["address"])
        if loc:
            db.execute("UPDATE packages SET lat=?,lon=? WHERE id=?", (loc[0], loc[1], r["id"]))
            db.commit()
            ok += 1
    db.close()
    flash(f"Geocodificación: {ok}/{len(rows)} resueltas.", "success")
    return redirect(url_for("packages_page", work_day_id=work_day_id))


@app.route("/admin/workdays/<int:work_day_id>/routes")
@login_required("admin")
def routes_page(work_day_id):
    db = get_db()
    wd = get_workday(db, work_day_id)
    drivers = db.execute(
        """
        SELECT d.*, COUNT(p.id) stops,
               COALESCE(SUM(p.lat IS NOT NULL AND p.lon IS NOT NULL),0) geocoded,
               COALESCE(SUM(p.status='delivered'),0) delivered
        FROM drivers d
        LEFT JOIN packages p ON p.driver_id=d.id AND p.work_day_id=?
        WHERE d.organization_id=? AND d.active=1
        GROUP BY d.id ORDER BY d.name
        """, (work_day_id, get_org_id())
    ).fetchall()
    runs = db.execute(
        """
        SELECT rr.*, d.name driver_name
        FROM route_runs rr JOIN drivers d ON d.id=rr.driver_id
        WHERE rr.work_day_id=? ORDER BY rr.id DESC LIMIT 100
        """, (work_day_id,)
    ).fetchall()
    google_ready = bool(os.environ.get("GOOGLE_CLOUD_PROJECT", "").strip()
                        and os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "").strip())
    db.close()
    return render_template(
        "routes.html", wd=wd, drivers=drivers, runs=runs, google_ready=google_ready
    )


@app.route("/admin/workdays/<int:work_day_id>/routes/optimize/<int:driver_id>", methods=["POST"])
@login_required("admin")
def optimize_route(work_day_id, driver_id):
    provider = request.form.get("provider", "local")
    if provider not in {"local", "google"}:
        provider = "local"
    db = get_db()
    wd = get_workday(db, work_day_id)
    org = db.execute("SELECT * FROM organizations WHERE id=?", (get_org_id(),)).fetchone()
    try:
        count, km, mins, provider_name, note = optimize_driver_route(
            db, org, wd, driver_id, provider
        )
        db.commit()
        flash(
            f"Ruta {provider_name}: {count} paradas · {km:.1f} km · {mins} min.",
            "success"
        )
    except Exception as exc:
        db.rollback()
        flash(f"No se pudo optimizar: {exc}", "error")
    finally:
        db.close()
    return redirect(url_for("routes_page", work_day_id=work_day_id))


@app.route("/admin/drivers")
@login_required("admin")
def drivers_page():
    db = get_db()
    drivers = db.execute(
        """
        SELECT d.*, u.username
        FROM drivers d LEFT JOIN users u ON u.driver_id=d.id
        WHERE d.organization_id=?
        ORDER BY d.name
        """, (get_org_id(),)
    ).fetchall()
    db.close()
    return render_template("drivers.html", drivers=drivers)


@app.route("/admin/drivers", methods=["POST"])
@login_required("admin")
def create_driver():
    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip().lower()
    try:
        pay = float(request.form.get("pay_per_delivery") or 0.85)
    except Exception:
        pay = 0.85
    if not name:
        flash("Nombre requerido.", "error")
        return redirect(url_for("drivers_page"))

    db = get_db()
    cur = db.execute(
        "INSERT INTO drivers(organization_id,name,email,pay_per_delivery,created_at) VALUES(?,?,?,?,?)",
        (get_org_id(), name, email, pay, now_iso())
    )
    driver_id = cur.lastrowid
    stem = "".join(c for c in name.lower().replace(" ", "") if c.isalnum()) or "driver"
    username = stem
    n = 1
    while db.execute(
        "SELECT 1 FROM users WHERE organization_id=? AND username=?",
        (get_org_id(), username)
    ).fetchone():
        n += 1
        username = f"{stem}{n}"
    db.execute(
        "INSERT INTO users(organization_id,email,username,password_hash,role,driver_id) VALUES(?,?,?,?,?,?)",
        (get_org_id(), email or None, username, generate_password_hash("1234"), "driver", driver_id)
    )
    db.commit()
    db.close()
    flash(f"Repartidor creado · usuario {username} · clave temporal 1234.", "success")
    return redirect(url_for("drivers_page"))


@app.route("/admin/workdays/<int:work_day_id>/map")
@login_required("admin")
def live_map(work_day_id):
    db = get_db()
    wd = get_workday(db, work_day_id)
    points = db.execute(
        """
        SELECT p.id,p.code,p.address,p.lat,p.lon,p.status,p.sequence,
               d.name driver_name,d.id driver_id
        FROM packages p LEFT JOIN drivers d ON d.id=p.driver_id
        WHERE p.work_day_id=? AND p.lat IS NOT NULL AND p.lon IS NOT NULL
        ORDER BY d.id,p.sequence,p.id
        """, (work_day_id,)
    ).fetchall()
    drivers = db.execute(
        "SELECT * FROM drivers WHERE organization_id=? AND active=1 ORDER BY name",
        (get_org_id(),)
    ).fetchall()
    db.close()
    return render_template(
        "map.html", wd=wd, drivers=drivers, points=[dict(x) for x in points]
    )


@app.route("/api/admin/workdays/<int:work_day_id>/tracking")
@login_required("admin")
def api_admin_tracking(work_day_id):
    db = get_db()
    get_workday(db, work_day_id)
    drivers = db.execute(
        "SELECT id,name FROM drivers WHERE organization_id=? AND active=1 ORDER BY name",
        (get_org_id(),)
    ).fetchall()
    output = []
    for d in drivers:
        rows = db.execute(
            """
            SELECT lat,lon,accuracy,speed,heading,captured_at
            FROM location_updates
            WHERE work_day_id=? AND driver_id=?
            ORDER BY id DESC LIMIT 120
            """, (work_day_id, d["id"])
        ).fetchall()
        rows = list(reversed([dict(r) for r in rows]))
        output.append({
            "driver_id": d["id"],
            "name": d["name"],
            "track": rows,
            "last": rows[-1] if rows else None
        })
    db.close()
    return jsonify({"ok": True, "drivers": output, "server_time": now_iso()})


@app.route("/admin/workdays/<int:work_day_id>/settlements")
@login_required("admin")
def settlements_page(work_day_id):
    db = get_db()
    wd = get_workday(db, work_day_id)
    drivers = db.execute(
        "SELECT * FROM drivers WHERE organization_id=? AND active=1 ORDER BY name",
        (get_org_id(),)
    ).fetchall()
    for drv in drivers:
        count = db.execute(
            """
            SELECT COUNT(*) c FROM packages
            WHERE work_day_id=? AND driver_id=? AND status='delivered'
            """, (work_day_id, drv["id"])
        ).fetchone()["c"]
        amount = count * drv["pay_per_delivery"]
        db.execute(
            """
            INSERT INTO settlements(
                organization_id,work_day_id,driver_id,delivered_count,amount,status
            ) VALUES(?,?,?,?,?,'pending')
            ON CONFLICT(work_day_id,driver_id) DO UPDATE SET
                delivered_count=excluded.delivered_count,
                amount=excluded.amount
            """, (get_org_id(), work_day_id, drv["id"], count, amount)
        )
    db.commit()
    settlements = db.execute(
        """
        SELECT s.*,d.name driver_name,d.pay_per_delivery
        FROM settlements s JOIN drivers d ON d.id=s.driver_id
        WHERE s.work_day_id=? ORDER BY d.name
        """, (work_day_id,)
    ).fetchall()
    total = sum(float(r["amount"] or 0) for r in settlements)
    db.close()
    return render_template(
        "settlements.html", wd=wd, settlements=settlements, total=total
    )


@app.route("/admin/settlements/<int:settlement_id>/paid", methods=["POST"])
@login_required("admin")
def mark_settlement_paid(settlement_id):
    db = get_db()
    row = db.execute(
        "SELECT * FROM settlements WHERE id=? AND organization_id=?",
        (settlement_id, get_org_id())
    ).fetchone()
    if not row:
        abort(404)
    db.execute(
        "UPDATE settlements SET status='paid',paid_at=? WHERE id=?",
        (now_iso(), settlement_id)
    )
    db.commit()
    db.close()
    return redirect(request.referrer or url_for("admin_dashboard"))


@app.route("/admin/workdays/<int:work_day_id>/settlements/export")
@login_required("admin")
def export_settlements(work_day_id):
    db = get_db()
    wd = get_workday(db, work_day_id)
    rows = db.execute(
        """
        SELECT d.name,s.delivered_count,d.pay_per_delivery,s.amount,s.status
        FROM settlements s JOIN drivers d ON d.id=s.driver_id
        WHERE s.work_day_id=? ORDER BY d.name
        """, (work_day_id,)
    ).fetchall()
    db.close()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Jornada","Fecha","Repartidor","Entregados","Pago por entrega","Total","Estado"])
    for r in rows:
        w.writerow([
            wd["name"], wd["work_date"], r["name"], r["delivered_count"],
            r["pay_per_delivery"], r["amount"], r["status"]
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=liquidacion_{wd['work_date']}_{work_day_id}.csv"}
    )


@app.route("/admin/reports")
@login_required("admin")
def reports_page():
    db = get_db()
    workdays = db.execute(
        """
        SELECT w.id,w.work_date,w.name,w.status,
               COUNT(p.id) total,
               COALESCE(SUM(p.status='delivered'),0) delivered,
               COALESCE(SUM(p.status='failed'),0) failed
        FROM work_days w LEFT JOIN packages p ON p.work_day_id=w.id
        WHERE w.organization_id=?
        GROUP BY w.id ORDER BY w.work_date DESC,w.id DESC LIMIT 60
        """, (get_org_id(),)
    ).fetchall()
    db.close()
    return render_template("reports.html", workdays=workdays)


@app.route("/admin/settings", methods=["GET", "POST"])
@login_required("admin")
def settings_page():
    db = get_db()
    org = db.execute("SELECT * FROM organizations WHERE id=?", (get_org_id(),)).fetchone()
    if request.method == "POST":
        name = request.form.get("name", "").strip() or org["name"]
        currency = request.form.get("currency", "").strip().upper() or "EUR"
        timezone_name = request.form.get("timezone", "").strip() or TIMEZONE_NAME
        depot_lat = request.form.get("depot_lat", type=float)
        depot_lon = request.form.get("depot_lon", type=float)
        db.execute(
            """
            UPDATE organizations
            SET name=?,currency=?,timezone=?,depot_lat=?,depot_lon=?
            WHERE id=?
            """, (name, currency, timezone_name, depot_lat, depot_lon, get_org_id())
        )
        db.commit()
        flash("Configuración guardada.", "success")
        org = db.execute("SELECT * FROM organizations WHERE id=?", (get_org_id(),)).fetchone()
    google_ready = bool(os.environ.get("GOOGLE_CLOUD_PROJECT", "").strip()
                        and os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "").strip())
    geocode_ready = bool(os.environ.get("GOOGLE_MAPS_API_KEY", "").strip())
    db.close()
    return render_template(
        "settings.html", org=org, google_ready=google_ready, geocode_ready=geocode_ready
    )


@app.route("/health")
@app.route("/healthz")
def health():
    db_ok = False
    try:
        db = get_db()
        db.execute("SELECT 1").fetchone()
        db.close()
        db_ok = True
    except Exception:
        db_ok = False
    return jsonify({"ok": db_ok, "version": "0.3.0", "database": database_backend()}), (200 if db_ok else 503)


@app.route("/driver")
@login_required("driver")
def driver_home():
    db = get_db()
    did = session["driver_id"]
    org_id = get_org_id()
    drv = db.execute("SELECT * FROM drivers WHERE id=? AND organization_id=?", (did, org_id)).fetchone()
    wd = active_workday_for_driver(db, org_id)
    stats = {"assigned": 0, "delivered": 0, "pending": 0, "failed": 0}
    latest = None
    if wd:
        stats = dict(db.execute(
            """
            SELECT COUNT(*) assigned,
                   COALESCE(SUM(status='delivered'),0) delivered,
                   COALESCE(SUM(status='pending'),0) pending,
                   COALESCE(SUM(status='failed'),0) failed
            FROM packages WHERE work_day_id=? AND driver_id=?
            """, (wd["id"], did)
        ).fetchone())
        latest = db.execute(
            """
            SELECT * FROM route_runs
            WHERE work_day_id=? AND driver_id=?
            ORDER BY id DESC LIMIT 1
            """, (wd["id"], did)
        ).fetchone()
    db.close()
    return render_template("driver_home.html", driver=drv, wd=wd, stats=stats, latest=latest)


@app.route("/driver/route")
@login_required("driver")
def driver_route():
    db = get_db()
    did = session["driver_id"]
    wd = active_workday_for_driver(db, get_org_id())
    if not wd:
        db.close()
        flash("No hay una jornada activa.", "error")
        return redirect(url_for("driver_home"))
    packages = db.execute(
        """
        SELECT * FROM packages
        WHERE work_day_id=? AND driver_id=? AND status!='delivered'
        ORDER BY CASE WHEN sequence IS NULL THEN 1 ELSE 0 END,sequence,id
        """, (wd["id"], did)
    ).fetchall()
    db.close()
    return render_template("driver_route.html", wd=wd, packages=packages)



@app.route("/driver/intake")
@login_required("driver")
def driver_intake():
    db = get_db()
    wd = active_workday_for_driver(db, get_org_id())
    if not wd:
        db.close()
        flash("No hay una jornada activa.", "error")
        return redirect(url_for("driver_home"))
    # Recover any queued OCR jobs after a Render restart or worker recycle.
    resume_intake_jobs(wd["id"], session["driver_id"])
    counts = intake_carrier_counts(db, wd["id"], session["driver_id"])
    recent = db.execute(
        """
        SELECT id,code,tracking_code,carrier,address,postal_code,intake_confidence,intake_status,sequence
        FROM packages WHERE work_day_id=? AND driver_id=? AND intake_source='camera'
        ORDER BY id DESC LIMIT 30
        """, (wd["id"], session["driver_id"])
    ).fetchall()
    ocr_ready = bool(os.environ.get("GOOGLE_VISION_API_KEY","").strip() or (os.environ.get("GOOGLE_CLOUD_PROJECT","").strip() and os.environ.get("GOOGLE_APPLICATION_CREDENTIALS","").strip()))
    geocode_ready = bool(os.environ.get("GOOGLE_MAPS_API_KEY", "").strip())
    ocr_debug_enabled = os.environ.get("INTAKE_OCR_DEBUG", "1").strip().lower() not in {"0","false","no"}
    db.close()
    return render_template("driver_intake.html", wd=wd, counts=counts, recent=recent, ocr_ready=ocr_ready, geocode_ready=geocode_ready, ocr_debug_enabled=ocr_debug_enabled)


@app.route("/api/driver/intake/capture", methods=["POST"])
@login_required("driver")
def api_driver_intake_capture():
    """Accept the photograph quickly and return immediately; OCR continues in background."""
    db = get_db()
    wd = active_workday_for_driver(db, get_org_id())
    if not wd:
        db.close()
        return jsonify({"ok": False, "error": "No hay jornada activa"}), 409

    image = request.files.get("image")
    if not image or not image.filename:
        db.close()
        return jsonify({"ok": False, "error": "Falta la imagen de la etiqueta"}), 422
    image_bytes = image.read()
    if not image_bytes:
        db.close()
        return jsonify({"ok": False, "error": "Imagen vacía"}), 422
    if len(image_bytes) > 8 * 1024 * 1024:
        db.close()
        return jsonify({"ok": False, "error": "Imagen demasiado grande"}), 413

    raw_codes_value = request.form.get("raw_codes") or "[]"
    try:
        raw_codes = json.loads(raw_codes_value)
        if not isinstance(raw_codes, list):
            raw_codes = []
    except Exception:
        raw_codes = []
    legacy_raw = (request.form.get("raw_code") or "").strip()
    if legacy_raw and legacy_raw not in raw_codes:
        raw_codes.insert(0, legacy_raw)
    raw_codes = [str(x).strip()[:120] for x in raw_codes if str(x).strip()][:12]
    carrier_hint = (request.form.get("carrier") or "").strip().lower()
    if carrier_hint not in {"", "imile", "ecoscooting", "tipsa", "agencia"}:
        carrier_hint = ""

    cur = db.execute(
        """
        INSERT INTO intake_jobs(
            organization_id,work_day_id,driver_id,status,raw_codes,carrier_hint,
            image_data,image_mime,image_size,created_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?)
        """,
        (
            get_org_id(), wd["id"], session["driver_id"], "queued",
            json.dumps(raw_codes, ensure_ascii=False), carrier_hint,
            image_bytes, image.mimetype or "image/jpeg", len(image_bytes), now_iso(),
        ),
    )
    job_id = cur.lastrowid
    db.commit()
    db.close()
    schedule_intake_job(job_id)
    return jsonify({"ok": True, "accepted": True, "job_id": job_id, "status": "queued"}), 202


@app.route("/api/driver/intake/jobs")
@login_required("driver")
def api_driver_intake_jobs():
    db = get_db()
    wd = active_workday_for_driver(db, get_org_id())
    if not wd:
        db.close()
        return jsonify({"ok": False, "error": "No hay jornada activa"}), 409
    resume_intake_jobs(wd["id"], session["driver_id"])
    rows = db.execute(
        """
        SELECT id,status,package_id,result_json,error_text,created_at,started_at,completed_at
        FROM intake_jobs
        WHERE work_day_id=? AND driver_id=?
        ORDER BY id DESC LIMIT 60
        """, (wd["id"], session["driver_id"])
    ).fetchall()
    stats_rows = db.execute(
        """
        SELECT status,COUNT(*) total FROM intake_jobs
        WHERE work_day_id=? AND driver_id=? AND status IN ('queued','processing')
        GROUP BY status
        """, (wd["id"], session["driver_id"])
    ).fetchall()
    stats = {"queued": 0, "processing": 0}
    for r in stats_rows:
        stats[r["status"]] = int(r["total"])
    counts = intake_carrier_counts(db, wd["id"], session["driver_id"])
    db.close()

    jobs = []
    for row in rows:
        result = {}
        if row["result_json"]:
            try:
                result = json.loads(row["result_json"])
            except Exception:
                result = {}
        item = {
            "id": row["id"], "status": row["status"], "package_id": row["package_id"],
            "error": row["error_text"] or "", "created_at": row["created_at"],
            "completed_at": row["completed_at"], "result": result,
        }
        debug = _get_ocr_debug(row["id"])
        if debug:
            item["ocr_debug"] = debug
        jobs.append(item)
    return jsonify({"ok": True, "jobs": jobs, "queue": stats, "counts": counts})


@app.route("/driver/intake/<int:package_id>/review", methods=["GET","POST"])

@login_required("driver")
def driver_intake_review(package_id):
    db = get_db()
    p = db.execute("SELECT * FROM packages WHERE id=? AND driver_id=? AND organization_id=?", (package_id, session["driver_id"], get_org_id())).fetchone()
    if not p:
        db.close(); abort(404)
    if request.method == "POST":
        carrier = request.form.get("carrier","unknown")
        recipient = request.form.get("recipient_name","").strip()
        address = request.form.get("address","").strip()
        postal = request.form.get("postal_code","").strip()
        zone = request.form.get("route_zone","").strip()
        route_code = request.form.get("route_code","").strip()
        tracking = request.form.get("tracking_code","").strip() or p["code"]
        lat = lon = None
        if address:
            loc = geocode_google(address)
            if loc: lat, lon = loc
        intake_status = "ready" if address and tracking and carrier in {"imile","ecoscooting","tipsa"} else "review"
        db.execute(
            """
            UPDATE packages SET carrier=?,recipient_name=?,address=?,postal_code=?,route_zone=?,route_code=?,
                tracking_code=?,barcode=?,lat=?,lon=?,intake_status=? WHERE id=?
            """,
            (carrier,recipient,address or "PENDIENTE DE REVISIÓN",postal,zone,route_code,tracking,tracking,lat,lon,intake_status,package_id)
        )
        db.commit()
        wd = active_workday_for_driver(db, get_org_id())
        if wd and lat is not None:
            reoptimize_after_intake(db, wd, session["driver_id"]); db.commit()
        db.close(); flash("Paquete actualizado.", "success")
        return redirect(url_for("driver_intake"))
    db.close()
    return render_template("driver_intake_review.html", p=p)

@app.route("/driver/scan")
@login_required("driver")
def driver_scan():
    db = get_db()
    wd = active_workday_for_driver(db, get_org_id())
    db.close()
    if not wd:
        flash("No hay una jornada activa.", "error")
        return redirect(url_for("driver_home"))
    return render_template("driver_scan.html", wd=wd)


@app.route("/driver/scan/lookup", methods=["POST"])
@login_required("driver")
def driver_scan_lookup():
    raw_code = (request.form.get("code") or "").strip()
    if not raw_code:
        flash("No se recibió ningún código.", "error")
        return redirect(url_for("driver_scan"))
    db = get_db()
    wd = active_workday_for_driver(db, get_org_id())
    if not wd:
        db.close()
        return redirect(url_for("driver_home"))
    p = db.execute(
        """
        SELECT * FROM packages
        WHERE work_day_id=? AND driver_id=? AND (code=? OR barcode=?)
        ORDER BY id LIMIT 1
        """, (wd["id"], session["driver_id"], raw_code, raw_code)
    ).fetchone()
    db.execute(
        """
        INSERT INTO scan_events(
            organization_id,work_day_id,driver_id,raw_code,package_id,scan_type,captured_at
        ) VALUES(?,?,?,?,?,?,?)
        """, (
            get_org_id(), wd["id"], session["driver_id"], raw_code,
            p["id"] if p else None, "lookup", now_iso()
        )
    )
    db.commit()
    db.close()
    if not p:
        flash("Código no encontrado en tu jornada.", "error")
        return redirect(url_for("driver_scan"))
    return redirect(url_for("driver_package", package_id=p["id"]))


@app.route("/driver/package/<int:package_id>")
@login_required("driver")
def driver_package(package_id):
    db = get_db()
    did = session["driver_id"]
    p = db.execute(
        """
        SELECT p.*,w.name workday_name,w.work_date
        FROM packages p JOIN work_days w ON w.id=p.work_day_id
        WHERE p.id=? AND p.driver_id=? AND p.organization_id=?
        """, (package_id, did, get_org_id())
    ).fetchone()
    db.close()
    if not p:
        flash("Paquete no disponible.", "error")
        return redirect(url_for("driver_route"))
    return render_template("driver_package.html", p=p)


@app.route("/driver/package/<int:package_id>/deliver", methods=["POST"])
@login_required("driver")
def deliver_package(package_id):
    did = session["driver_id"]
    photo = request.files.get("photo")
    proof_bytes = None
    proof_mime = None
    proof_filename = None
    if photo and photo.filename:
        ext = photo.filename.rsplit(".", 1)[-1].lower() if "." in photo.filename else ""
        if ext not in ALLOWED_IMAGE_EXT:
            flash("La evidencia debe ser PNG, JPG, JPEG o WEBP.", "error")
            return redirect(url_for("driver_package", package_id=package_id))
        try:
            raw = photo.read()
            if len(raw) > 12 * 1024 * 1024:
                raise ValueError("archivo demasiado grande")
            img = Image.open(io.BytesIO(raw))
            img = ImageOps.exif_transpose(img).convert("RGB")
            img.thumbnail((1280, 1280))
            out = io.BytesIO()
            img.save(out, format="JPEG", quality=72, optimize=True)
            proof_bytes = out.getvalue()
            proof_mime = "image/jpeg"
            proof_filename = secure_filename(photo.filename)[:160] or f"evidencia_{package_id}.jpg"
        except Exception:
            flash("No se pudo procesar la foto de evidencia.", "error")
            return redirect(url_for("driver_package", package_id=package_id))

    lat = request.form.get("lat", type=float)
    lon = request.form.get("lon", type=float)
    accuracy = request.form.get("accuracy", type=float)
    notes = request.form.get("notes", "").strip()
    db = get_db()
    p = db.execute(
        "SELECT * FROM packages WHERE id=? AND driver_id=? AND organization_id=?",
        (package_id, did, get_org_id())
    ).fetchone()
    if not p:
        db.close()
        abort(404)
    db.execute(
        """
        UPDATE packages
        SET status='delivered',delivered_at=?,failure_reason=NULL,
            proof_image=COALESCE(?,proof_image),proof_mime=COALESCE(?,proof_mime),
            proof_filename=COALESCE(?,proof_filename),delivery_lat=?,delivery_lon=?,
            delivery_accuracy=?,notes=?,delivered_by_driver_id=?
        WHERE id=?
        """, (now_iso(), proof_bytes, proof_mime, proof_filename, lat, lon, accuracy, notes, did, package_id)
    )
    db.commit()
    db.close()
    flash("Entrega registrada.", "success")
    return redirect(url_for("driver_route"))


@app.route("/driver/package/<int:package_id>/fail", methods=["POST"])
@login_required("driver")
def fail_package(package_id):
    did = session["driver_id"]
    reason = request.form.get("reason", "Otro")
    notes = request.form.get("notes", "").strip()
    db = get_db()
    db.execute(
        """
        UPDATE packages SET status='failed',failure_reason=?,notes=?
        WHERE id=? AND driver_id=? AND organization_id=?
        """, (reason, notes, package_id, did, get_org_id())
    )
    db.commit()
    db.close()
    flash("Incidencia registrada.", "success")
    return redirect(url_for("driver_route"))


@app.route("/api/driver/location", methods=["POST"])
@login_required("driver")
def api_driver_location():
    data = request.get_json(silent=True) or {}
    try:
        lat = float(data["lat"])
        lon = float(data["lon"])
    except Exception:
        return jsonify({"ok": False, "error": "lat/lon invalid"}), 400

    db = get_db()
    wd = active_workday_for_driver(db, get_org_id())
    if not wd:
        db.close()
        return jsonify({"ok": False, "error": "no active workday"}), 409

    def num(name):
        try:
            value = data.get(name)
            return float(value) if value is not None else None
        except Exception:
            return None

    db.execute(
        """
        INSERT INTO location_updates(
            organization_id,work_day_id,driver_id,lat,lon,accuracy,speed,heading,captured_at
        ) VALUES(?,?,?,?,?,?,?,?,?)
        """,
        (
            get_org_id(), wd["id"], session["driver_id"], lat, lon,
            num("accuracy"), num("speed"), num("heading"), now_iso()
        )
    )
    db.commit()
    db.close()
    return jsonify({"ok": True, "captured_at": now_iso()})


@app.route("/proof/<int:package_id>")
@login_required()
def proof_image(package_id):
    db = get_db()
    p = db.execute("SELECT organization_id,proof_image,proof_mime FROM packages WHERE id=?", (package_id,)).fetchone()
    db.close()
    if not p or p["organization_id"] != get_org_id() or not p["proof_image"]:
        abort(404)
    return Response(bytes(p["proof_image"]), mimetype=p["proof_mime"] or "image/jpeg", headers={"Cache-Control":"private, max-age=3600"})


@app.route("/uploads/<path:name>")
@login_required()
def uploads(name):
    # Legacy local evidence route kept only for older pilot data.
    return send_from_directory(UPLOAD_DIR, name)


@app.route("/manifest.webmanifest")
def manifest():
    return send_from_directory(BASE_DIR / "static", "manifest.webmanifest")


@app.route("/sw.js")
def service_worker():
    return send_from_directory(BASE_DIR / "static", "sw.js", mimetype="application/javascript")


@app.errorhandler(413)
def too_large(_e):
    flash("Archivo demasiado grande (máximo 12 MB).", "error")
    return redirect(request.referrer or url_for("home"))


if __name__ == "__main__":
    init_db()
    print("\\nRouteOps V0.3.1.4 Intelligent OCR Pipeline local listo en http://127.0.0.1:5000")
    print("Admin: admin@routeops.local / demo123")
    print("Repartidor: carlos / 1234")
    print(f"Zona horaria: {TIMEZONE_NAME}\\n")
    app.run(host="127.0.0.1", port=5000, debug=False)
